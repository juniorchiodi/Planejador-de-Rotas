import pandas as pd
from fpdf import FPDF
from geopy.distance import geodesic
import numpy as np
import os
import unicodedata
import json
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from tqdm import tqdm
import colorama
from colorama import Fore, Style
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import requests
import openpyxl
from openpyxl.styles import PatternFill
import time  # Corrigido: necessário para time.sleep
import re

# Inicializa colorama para Windows
colorama.init()

def print_colorido(texto, cor=Fore.WHITE, estilo=Style.NORMAL):
    print(f"{estilo}{cor}{texto}{Style.RESET_ALL}")

# Função para detectar se o endereço está no formato de coordenadas

def is_coordenada(texto):
    if not isinstance(texto, str):
        return False
    # Regex para latitude e longitude: -22.188655, -48.615678 (com ou sem texto adicional)
    padrao = r"^\s*(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)(?:\s*[,;]\s*.*)?$"
    return re.match(padrao, texto) is not None

def extrair_coordenada(texto):
    # Regex para extrair apenas latitude e longitude, ignorando texto adicional
    padrao = r"^\s*(-?\d{1,2}\.\d+),\s*(-?\d{1,3}\.\d+)"
    m = re.match(padrao, texto)
    if m:
        return (float(m.group(1)), float(m.group(2)))
    return None

# Função para remover acentos
def remover_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                  if unicodedata.category(c) != 'Mn')

# Cache para geocodificação com timestamp
CACHE_FILE = "geocodificacao_cache.json"
CACHE_EXPIRATION_DAYS = 30  # Cache expira após 30 dias

def carregar_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
                current_time = datetime.now()
                cache_data = {
                    k: v for k, v in cache_data.items()
                    if datetime.fromisoformat(v['timestamp']) + timedelta(days=CACHE_EXPIRATION_DAYS) > current_time
                }
                return cache_data
        except Exception as e:
            print_colorido(f"Erro ao carregar cache: {str(e)}", Fore.RED)
            return {}
    return {}

def salvar_cache(cache):
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print_colorido(f"Erro ao salvar cache: {str(e)}", Fore.RED)

def expandir_abreviacoes(endereco):
    abreviacoes = {
        'Av.': 'Avenida',
        'Av ': 'Avenida',
        'R.': 'Rua',
        'R ': 'Rua',
        'Al.': 'Alameda',
        'Al ': 'Alameda',
        'Tv.': 'Travessa',
        'Tv ': 'Travessa',
        'Est.': 'Estrada',
        'Est ': 'Estrada'
    }
    for abrev, completo in abreviacoes.items():
        endereco = endereco.replace(abrev, completo)
    return endereco

def geocodificar_endereco(endereco, max_tentativas=3):
    endereco_sem_acento = remover_acentos(endereco)
    endereco_formatado = f"{endereco_sem_acento}, Brasil"
    url = f"https://nominatim.openstreetmap.org/search?q={requests.utils.quote(endereco_formatado)}&format=json&limit=1"
    headers = {
        'User-Agent': 'RotaEntregas/1.0 (https://github.com/juniorchiodi/Planejador-de-Rotas; juninho.junirj@gmail.com) Python/3.x'
    }
    for tentativa in range(max_tentativas):
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 403:
                time.sleep(10)
                continue
            response.raise_for_status()
            data = response.json()
            if data:
                lat = float(data[0]['lat'])
                lon = float(data[0]['lon'])
                time.sleep(1)
                return {
                    'coords': (lat, lon),
                    'timestamp': datetime.now().isoformat()
                }
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if tentativa < max_tentativas - 1:
                time.sleep((tentativa + 1) * 5)
            continue
        except Exception:
            if tentativa < max_tentativas - 1:
                time.sleep(5)
            continue
    return None

def calcular_distancia_osrm(coords1, coords2, max_tentativas=3):
    try:
        lon1, lat1 = coords1[1], coords1[0]
        lon2, lat2 = coords2[1], coords2[0]
        url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}?overview=full&alternatives=true&steps=true&annotations=distance"
        for tentativa in range(max_tentativas):
            try:
                session = requests.Session()
                retry = Retry(total=3, backoff_factor=0.5)
                adapter = HTTPAdapter(max_retries=retry)
                session.mount('http://', adapter)
                session.mount('https://', adapter)
                response = session.get(url, timeout=15)
                if response.status_code == 200:
                    data = response.json()
                    if data['code'] == 'Ok':
                        rotas = data['routes']
                        if rotas:
                            rotas_ordenadas = sorted(rotas, key=lambda x: x['distance'])
                            melhor_rota = rotas_ordenadas[0]
                            distancia_total = 0
                            for leg in melhor_rota['legs']:
                                for step in leg['steps']:
                                    distancia_total += step['distance']
                            return distancia_total / 1000
                elif response.status_code == 429:
                    print_colorido("⚠️ Rate limit atingido. Aguardando 5 segundos...", Fore.YELLOW)
                    time.sleep(5)
                    continue
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                print_colorido(f"⚠️ Erro de conexão na tentativa {tentativa + 1}: {str(e)}", Fore.RED)
                if tentativa < max_tentativas - 1:
                    tempo_espera = (tentativa + 1) * 5
                    print_colorido(f"   Aguardando {tempo_espera} segundos...", Fore.YELLOW)
                    time.sleep(tempo_espera)
                continue
            except Exception as e:
                print_colorido(f"❌ Erro inesperado: {str(e)}", Fore.RED)
                if tentativa < max_tentativas - 1:
                    print_colorido("   Aguardando 5 segundos...", Fore.YELLOW)
                    time.sleep(5)
                continue
    except Exception as e:
        print_colorido(f"Erro ao calcular distância OSRM: {str(e)}", Fore.RED)
    return None

def calcular_distancia_rua(coords1, coords2):
    try:
        lat1, lon1 = float(coords1[0]), float(coords1[1])
        lat2, lon2 = float(coords2[0]), float(coords2[1])
        if not (-90 <= lat1 <= 90) or not (-90 <= lat2 <= 90) or \
           not (-180 <= lon1 <= 180) or not (-180 <= lon2 <= 180):
            print_colorido(f"Coordenadas inválidas: ({lat1}, {lon1}) ou ({lat2}, {lon2})", Fore.RED)
            return float('inf')
        dist = geodesic((lat1, lon1), (lat2, lon2)).kilometers
        if dist > 500:
            print_colorido(f"Distância suspeita: {dist:.2f}km entre ({lat1}, {lon1}) e ({lat2}, {lon2})", Fore.YELLOW)
            return float('inf')
        return dist
    except Exception as e:
        print_colorido(f"Erro ao calcular distância: {str(e)}", Fore.RED)
        return float('inf')

def calcular_distancia_final(coords1, coords2):
    dist_osrm = calcular_distancia_osrm(coords1, coords2)
    if dist_osrm is not None:
        return round(dist_osrm * 1.1, 1)
    dist_geodesica = calcular_distancia_rua(coords1, coords2)
    return round(dist_geodesica * 1.15, 1)

DISTANCE_CACHE_FILE = "distance_cache.json"
DISTANCE_CACHE_EXPIRATION_DAYS = 30

def carregar_cache_distancia():
    if os.path.exists(DISTANCE_CACHE_FILE):
        try:
            with open(DISTANCE_CACHE_FILE, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
                current_time = datetime.now()
                cache_data = {
                    k: v for k, v in cache_data.items()
                    if datetime.fromisoformat(v['timestamp']) + timedelta(days=DISTANCE_CACHE_EXPIRATION_DAYS) > current_time
                }
                return cache_data
        except Exception as e:
            print_colorido(f"Erro ao carregar cache de distância: {str(e)}", Fore.RED)
            return {}
    return {}

def salvar_cache_distancia(cache):
    try:
        with open(DISTANCE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print_colorido(f"Erro ao salvar cache de distância: {str(e)}", Fore.RED)

def calcular_distancia_com_cache(coords1, coords2):
    key = f"{coords1[0]},{coords1[1]}_{coords2[0]},{coords2[1]}"
    cache = carregar_cache_distancia()
    if key in cache:
        distancia_cache = cache[key]['distance']
        distancia_geodesica = calcular_distancia_rua(coords1, coords2)
        if distancia_geodesica > 0 and distancia_cache > distancia_geodesica * 2:
            print_colorido(f"⚠️ Distância no cache muito maior que a geodésica. Recalculando...", Fore.YELLOW)
            distancia = calcular_distancia_final(coords1, coords2)
        else:
            return distancia_cache
    distancia = calcular_distancia_final(coords1, coords2)
    cache[key] = {
        'distance': distancia,
        'timestamp': datetime.now().isoformat()
    }
    salvar_cache_distancia(cache)
    return distancia

def identificar_outliers(dist_matrix, enderecos_validos, limite_desvio=2):
    """Identifica pontos que estão muito distantes da média"""
    n = len(dist_matrix)
    if n <= 1:
        return [], []

    # Calcula a média e desvio padrão das distâncias
    distancias = []
    for i in range(n):
        for j in range(n):
            if i != j and dist_matrix[i][j] != float('inf'):
                distancias.append(dist_matrix[i][j])
    
    if not distancias:
        return [], []

    media = sum(distancias) / len(distancias)
    desvio = (sum((x - media) ** 2 for x in distancias) / len(distancias)) ** 0.5
    
    # Identifica outliers
    outliers = []
    pontos_principais = []
    
    for i in range(n):
        distancias_ponto = [dist_matrix[i][j] for j in range(n) if i != j and dist_matrix[i][j] != float('inf')]
        if not distancias_ponto:
            continue
            
        media_ponto = sum(distancias_ponto) / len(distancias_ponto)
        if media_ponto > media + (limite_desvio * desvio):
            outliers.append(i)
            print_colorido(f"Ponto identificado como outlier: {enderecos_validos[i]} (distância média: {media_ponto:.2f} km)", Fore.YELLOW)
        else:
            pontos_principais.append(i)
    
    return pontos_principais, outliers

def encontrar_melhor_rota(dist_matrix, enderecos_validos):
    """Encontra a melhor rota sempre indo para o vizinho mais próximo"""
    n = len(dist_matrix)
    if n <= 1:
        return [0]

    # Função para verificar se a distância é aceitável
    def distancia_aceitavel(dist):
        return dist < 20  # Reduzido para 20km para priorizar pontos muito próximos

    # Função para encontrar o próximo ponto mais próximo com validação
    def encontrar_proximo_ponto(ponto_atual, pontos_nao_visitados):
        # Primeiro tenta encontrar um ponto muito próximo (até 5km)
        pontos_muito_proximos = [p for p in pontos_nao_visitados 
                               if dist_matrix[ponto_atual][p] < 5]
        if pontos_muito_proximos:
            return min(pontos_muito_proximos, key=lambda x: dist_matrix[ponto_atual][x])
        
        # Depois tenta encontrar um ponto próximo (até 10km)
        pontos_proximos = [p for p in pontos_nao_visitados 
                         if dist_matrix[ponto_atual][p] < 10]
        if pontos_proximos:
            return min(pontos_proximos, key=lambda x: dist_matrix[ponto_atual][x])
        
        # Depois tenta encontrar um ponto com distância aceitável (até 20km)
        pontos_validos = [p for p in pontos_nao_visitados 
                         if distancia_aceitavel(dist_matrix[ponto_atual][p])]
        if pontos_validos:
            return min(pontos_validos, key=lambda x: dist_matrix[ponto_atual][x])
        
        # Se não encontrar pontos próximos, usa o mais próximo disponível
        return min(pontos_nao_visitados, key=lambda x: dist_matrix[ponto_atual][x])

    # Começa do ponto de partida (índice 0)
    rota = [0]
    pontos_nao_visitados = set(range(1, n))
    
    # Enquanto houver pontos não visitados
    while pontos_nao_visitados:
        ponto_atual = rota[-1]
        
        # Encontra o próximo ponto mais próximo com validação
        proximo_ponto = encontrar_proximo_ponto(ponto_atual, pontos_nao_visitados)
        
        # Adiciona o ponto à rota
        rota.append(proximo_ponto)
        pontos_nao_visitados.remove(proximo_ponto)
        
        # Mostra a distância para o próximo ponto
        distancia = dist_matrix[ponto_atual][proximo_ponto]
        print_colorido(f"De {enderecos_validos[ponto_atual]} para {enderecos_validos[proximo_ponto]}: {distancia:.2f} km", Fore.WHITE)
    
    # Tenta otimizar a rota verificando se há pontos que podem ser reordenados
    rota_otimizada = rota.copy()
    melhorou = True
    
    while melhorou:
        melhorou = False
        for i in range(1, len(rota_otimizada) - 1):
            # Verifica se trocar a ordem de dois pontos melhora a distância total
            dist_atual = (dist_matrix[rota_otimizada[i-1]][rota_otimizada[i]] + 
                         dist_matrix[rota_otimizada[i]][rota_otimizada[i+1]])
            dist_nova = (dist_matrix[rota_otimizada[i-1]][rota_otimizada[i+1]] + 
                        dist_matrix[rota_otimizada[i+1]][rota_otimizada[i]])
            
            if dist_nova < dist_atual:
                # Troca os pontos
                rota_otimizada[i], rota_otimizada[i+1] = rota_otimizada[i+1], rota_otimizada[i]
                melhorou = True
                print_colorido(f"Otimização: Reordenando pontos {i} e {i+1}", Fore.YELLOW)
    
    # Calcula a distância total da rota otimizada
    distancia_total = 0
    for i in range(len(rota_otimizada) - 1):
        distancia_total += dist_matrix[rota_otimizada[i]][rota_otimizada[i + 1]]
    
    print_colorido(f"\nDistância total da rota: {distancia_total:.2f} km", Fore.GREEN)
    return rota_otimizada

# CONFIGURAÇÕES
arquivo_excel = "ENDERECOS-ROTA.xlsx"
nome_coluna_enderecos = "Endereco"
nome_coluna_nomes = "Nome"
ponto_partida = "Rua Floriano Peixoto, 368, Centro, Itapuí - SP"

# Solicitar a cidade ao usuário
cidade = input("Digite a cidade das entregas: ").strip()

try:
    print_colorido("\n🚀 Iniciando processamento...", Fore.GREEN, Style.BRIGHT)
    
    # Verificar se o arquivo Excel existe
    if not os.path.exists(arquivo_excel):
        print_colorido(f"❌ Erro: O arquivo {arquivo_excel} não foi encontrado.", Fore.RED)
        exit(1)

    # LER PLANILHA
    print_colorido("\n📊 Lendo planilha...", Fore.CYAN)
    try:
        df = pd.read_excel(arquivo_excel)
        df = df[df[nome_coluna_nomes].notna() & (df[nome_coluna_nomes] != "")]
        enderecos = df[nome_coluna_enderecos].dropna().tolist()
        nomes = df[nome_coluna_nomes].fillna("").tolist()
        print_colorido(f"✅ Total de endereços encontrados: {len(enderecos)}", Fore.GREEN)
    except Exception as e:
        print_colorido(f"❌ Erro ao ler planilha: {str(e)}", Fore.RED)
        exit(1)
    
    if not enderecos:
        print_colorido("❌ Erro: Nenhum endereço encontrado na planilha.", Fore.RED)
        exit(1)

    # GEOCODIFICAÇÃO
    print_colorido("\n🌍 Iniciando geocodificação...", Fore.CYAN)
    coordenadas = []
    enderecos_validos = []
    enderecos_com_erro = []

    # Primeiro, geocodificar o ponto de partida
    print_colorido(f"\n📍 Processando ponto de partida: {ponto_partida}", Fore.CYAN)
    cache = carregar_cache()
    if is_coordenada(ponto_partida):
        coords = extrair_coordenada(ponto_partida)
        if coords:
            coordenadas.append(coords)
            enderecos_validos.append(ponto_partida)
        else:
            print_colorido(f"❌ Erro: Não foi possível interpretar as coordenadas do ponto de partida: {ponto_partida}", Fore.RED)
            exit(1)
    elif ponto_partida in cache:
        print_colorido("✅ Usando coordenadas do cache para ponto de partida", Fore.GREEN)
        coordenadas.append(cache[ponto_partida]['coords'])
        enderecos_validos.append(ponto_partida)
    else:
        try:
            resultado = geocodificar_endereco(ponto_partida)
            if resultado:
                coordenadas.append(resultado['coords'])
                enderecos_validos.append(ponto_partida)
                cache[ponto_partida] = resultado
                salvar_cache(cache)
            else:
                print_colorido(f"❌ Erro: Não foi possível geocodificar o ponto de partida: {ponto_partida}", Fore.RED)
                exit(1)
        except Exception as e:
            print_colorido(f"❌ Erro ao geocodificar ponto de partida: {str(e)}", Fore.RED)
            exit(1)

    # Função para processar endereços em paralelo
    def processar_endereco(endereco):
        cache = carregar_cache()
        if is_coordenada(endereco):
            coords = extrair_coordenada(endereco)
            if coords:
                return endereco, coords, 'coordenada'
            else:
                return endereco, None, 'erro'
        if endereco in cache:
            # Não imprime aqui, apenas retorna status
            return endereco, cache[endereco]['coords'], 'cache'
        resultado = geocodificar_endereco(endereco)
        if resultado:
            cache[endereco] = resultado
            salvar_cache(cache)
            return endereco, resultado['coords'], 'geocodificado'
        return endereco, None, 'erro'

    # Processar endereços em paralelo com mais workers
    print_colorido("\n🔄 Geocodificando endereços...", Fore.CYAN)
    with ThreadPoolExecutor(max_workers=6) as executor:  # Reduzido para 6 workers para maior estabilidade
        resultados = list(tqdm(executor.map(processar_endereco, enderecos), 
                             total=len(enderecos),
                             desc="Progresso",
                             unit="endereço"))
    
    # Filtrar resultados válidos e coletar erros
    for i, (endereco, coords, status) in enumerate(resultados, 1):
        if coords:
            coordenadas.append(coords)
            enderecos_validos.append(endereco)
        else:
            enderecos_com_erro.append((i, endereco))
        # Imprime o status de cada endereço em ordem
        if status == 'cache':
            print_colorido(f"Usando cache para: {endereco}", Fore.YELLOW)
        elif status == 'geocodificado':
            print_colorido(f"Geocodificado: {endereco}", Fore.GREEN)
        elif status == 'coordenada':
            print_colorido(f"Endereço já é coordenada: {endereco}", Fore.CYAN)
        else:
            print_colorido(f"Erro ao geocodificar: {endereco}", Fore.RED)

    # NOVO: Marcar células dos endereços com erro em vermelho na planilha
    def marcar_enderecos_erro_excel(arquivo_excel, nome_coluna_enderecos, enderecos_com_erro):
        try:
            wb = openpyxl.load_workbook(arquivo_excel)
            ws = wb.active
            # Encontrar o número de colunas
            num_cols = ws.max_column
            # Preencher de vermelho todas as células da linha dos endereços com erro
            fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            for linha, endereco in enderecos_com_erro:
                for col in range(1, num_cols + 1):
                    ws.cell(row=linha+1, column=col).fill = fill  # +1 por causa do header
            wb.save(arquivo_excel)
            print_colorido(f"Linhas dos endereços com erro marcadas em vermelho na planilha.", Fore.RED)
        except Exception as e:
            print_colorido(f"Erro ao marcar células na planilha: {str(e)}", Fore.RED)

    if enderecos_com_erro:
        marcar_enderecos_erro_excel(arquivo_excel, nome_coluna_enderecos, enderecos_com_erro)

    if len(enderecos_validos) <= 1:
        print_colorido("❌ Erro: Nenhum endereço foi geocodificado com sucesso além do ponto de partida.", Fore.RED)
        exit(1)

    print_colorido(f"\n✅ Total de endereços geocodificados com sucesso: {len(enderecos_validos)}", Fore.GREEN)
    print_colorido(f"⚠️ Total de endereços com erro: {len(enderecos_com_erro)}", Fore.YELLOW)

    # MATRIZ DE DISTÂNCIA
    print_colorido("\n📏 Calculando matriz de distância...", Fore.CYAN)
    n = len(coordenadas)
    dist_matrix = np.zeros((n, n))

    # Função para calcular distâncias em paralelo
    def calcular_distancia_paralela(args):
        i, j, coords1, coords2 = args
        if i != j:
            dist = calcular_distancia_com_cache(coords1, coords2)
            return i, j, dist
        return i, j, 0

    # Criar lista de argumentos para processamento paralelo
    args_list = []
    for i in range(n):
        for j in range(n):
            args_list.append((i, j, coordenadas[i], coordenadas[j]))

    # Calcular distâncias em paralelo
    with ThreadPoolExecutor(max_workers=6) as executor:  # Reduzido para 6 workers para maior estabilidade
        resultados = list(tqdm(executor.map(calcular_distancia_paralela, args_list),
                             total=len(args_list),
                             desc="Calculando distâncias",
                             unit="par"))

    # Preencher matriz de distância com resultados
    for i, j, dist in resultados:
        dist_matrix[i][j] = dist
        if dist != float('inf'):
            print_colorido(f"   De {enderecos_validos[i]} para {enderecos_validos[j]}: {dist:.2f} km", Fore.WHITE)

    # ENCONTRAR MELHOR ROTA
    print_colorido("\n🗺️ Calculando melhor rota...", Fore.CYAN)
    ordem_rota = encontrar_melhor_rota(dist_matrix, enderecos_validos)
    
    if ordem_rota is None:
        print_colorido("❌ Erro: Não foi possível encontrar uma rota válida", Fore.RED)
        exit(1)
    
    # Verifica se a rota está correta
    if len(ordem_rota) != len(coordenadas):
        print_colorido("❌ Erro: A rota não inclui todos os pontos!", Fore.RED)
        exit(1)

    # Calcula a distância total da rota e as distâncias parciais
    distancia_total = 0
    distancias_parciais = []
    for i in range(len(ordem_rota) - 1):
        dist = dist_matrix[ordem_rota[i]][ordem_rota[i + 1]]
        distancia_total += dist
        distancias_parciais.append(dist)
    print_colorido(f"\n📊 Distância total da rota: {distancia_total:.2f} km", Fore.GREEN)

    enderecos_ordenados = [enderecos_validos[i] for i in ordem_rota]
    nomes_ordenados = [nomes[enderecos.index(endereco)] if endereco in enderecos else "" for endereco in enderecos_ordenados]
    links = [f"https://www.google.com/maps/search/?api=1&query={remover_acentos(e).replace(' ', '+')}" for e in enderecos_ordenados]

    # GERAR PDF
    print_colorido("\n📄 Gerando PDF...", Fore.CYAN)
    pdf = FPDF()
    pdf.add_page()
    
    # Adicionar logo
    if os.path.exists("./assets/logo.png"):
        pdf.image("./assets/logo.png", x=170, y=10, w=31.5)
    
    pdf.set_font("Arial", "B", 16)
    data_atual = datetime.now().strftime("%Y/%m/%d")
    titulo = f"{data_atual} - Rota de Entregas - {cidade}"
    nome_arquivo = remover_acentos(titulo).replace(" ", "_").replace("/", "-")
    pasta_rotas = "ROTAS-GERADAS"
    if not os.path.exists(pasta_rotas):
        os.makedirs(pasta_rotas)
    arquivo_saida_pdf = os.path.join(pasta_rotas, f"{nome_arquivo}.pdf")
    pdf.cell(0, 10, titulo, ln=True, align="C")
    pdf.ln(10)

    # Informações gerais
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Informações Gerais:", ln=True)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Ponto de Partida: {ponto_partida}", ln=True)
    pdf.cell(0, 10, f"Distância Total Estimada: {distancia_total:.2f} km", ln=True)
    pdf.cell(0, 10, f"Número de Entregas: {len(enderecos_ordenados) - 1}", ln=True)
    pdf.cell(0, 10, f"Total de Endereços com Erro: {len(enderecos_com_erro)}", ln=True)
    pdf.ln(10)

    # Cabeçalho da tabela
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(255, 0, 0)  # Vermelho
    pdf.set_text_color(255, 255, 255)  # Texto branco
    pdf.cell(10, 10, "Nº", 1, 0, "C", True)
    pdf.cell(45, 10, "Nome", 1, 0, "C", True)
    pdf.cell(65, 10, "Endereço", 1, 0, "C", True)
    pdf.cell(30, 10, "Distância", 1, 0, "C", True)
    pdf.cell(40, 10, "Link", 1, 1, "C", True)

    # Dados da tabela
    pdf.set_font("Arial", "", 10)
    pdf.set_text_color(0, 0, 0)  # Texto preto
    pdf.set_fill_color(255, 240, 240)
    for i, (nome, endereco, link, dist) in enumerate(zip(nomes_ordenados[1:], enderecos_ordenados[1:], links[1:], distancias_parciais), 1):
        if pdf.get_y() > 250:
            pdf.add_page()
            pdf.set_font("Arial", "B", 12)
            pdf.set_fill_color(255, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(10, 10, "Nº", 1, 0, "C", True)
            pdf.cell(45, 10, "Nome", 1, 0, "C", True)
            pdf.cell(65, 10, "Endereço", 1, 0, "C", True)
            pdf.cell(30, 10, "Distância", 1, 0, "C", True)
            pdf.cell(40, 10, "Link", 1, 1, "C", True)
            pdf.set_font("Arial", "", 10)
            pdf.set_text_color(0, 0, 0)
            pdf.set_fill_color(255, 240, 240)
        pdf.cell(10, 15, '#' + str(i), 1, 0, "C", True)
        pdf.cell(45, 15, str(nome)[:30], 1, 0, "L", True)
        pdf.cell(65, 15, str(endereco)[:45], 1, 0, "L", True)
        pdf.cell(30, 15, f"{round(dist, 1)} km", 1, 0, "C", True)
        pdf.set_text_color(255, 0, 0)
        pdf.cell(40, 15, "Ver no Maps", 1, 1, "C", True, link=link)
        pdf.set_text_color(0, 0, 0)

    # Adicionar seção de endereços com erro
    if enderecos_com_erro:
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Endereços com Erro na Geocodificação", ln=True, align="C")
        pdf.ln(10)
        pdf.set_font("Arial", "B", 12)
        pdf.set_fill_color(255, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(20, 10, "Linha", 1, 0, "C", True)
        pdf.cell(50, 10, "Nome", 1, 0, "C", True)
        pdf.cell(120, 10, "Endereço", 1, 1, "C", True)
        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(0, 0, 0)
        pdf.set_fill_color(255, 240, 240)
        for linha, endereco in enderecos_com_erro:
            if pdf.get_y() > 250:
                pdf.add_page()
                pdf.set_font("Arial", "B", 12)
                pdf.set_fill_color(255, 0, 0)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(20, 10, "Linha", 1, 0, "C", True)
                pdf.cell(50, 10, "Nome", 1, 0, "C", True)
                pdf.cell(120, 10, "Endereço", 1, 1, "C", True)
                pdf.set_font("Arial", "", 10)
                pdf.set_text_color(0, 0, 0)
                pdf.set_fill_color(255, 240, 240)
            # Obter nome do cliente de forma segura
            nome_cliente = ""
            if 0 <= linha - 1 < len(nomes):
                nome_cliente = nomes[linha - 1]
            pdf.cell(20, 15, str(linha), 1, 0, "C", True)
            pdf.cell(50, 15, str(nome_cliente)[:30], 1, 0, "L", True)
            pdf.cell(120, 15, str(endereco)[:70], 1, 1, "L", True)

    pdf.output(arquivo_saida_pdf)
    print_colorido(f"\n✅ PDF gerado com sucesso: {arquivo_saida_pdf}", Fore.GREEN)

except Exception as e:
    print_colorido(f"\n❌ Erro inesperado: {str(e)}", Fore.RED)
    import traceback
    print_colorido("Detalhes do erro:", Fore.RED)
    print_colorido(traceback.format_exc(), Fore.RED)
    exit(1)
